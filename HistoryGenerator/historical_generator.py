import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import calendar
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# --- CONFIGURATION ---
# Read configuration from Data.xlsx (same folder as script)
config_file = "Data.xlsx"
master_df = pd.read_excel(config_file)

# Convert to dictionary for easy lookup
items = {}
for _, row in master_df.iterrows():
    items[row["Item Code"]] = {
        "name": row["Item Name"],
        "avg_daily": row["Avg Daily Use"],
        "price": row["Price (€)"],
        "lead_time": row["Lead Time(days)"],
        "moq": row["MOQ"],
        "planning_time_fence": row["Planning Time Fence(days)"],
        "parent_item_code": row["Parent Item Code"],
        "waste_pct": row["Avg Waste(%)"],   # decimal fraction (0.04 = 4%)
        "max_stock": row["Max Stock"],
        "reorder_point": row["Reorder Point"]
    }

# Serbian public holidays (2025–2026)
holidays = {
    "2025-01-01": {"event": "New Year’s Day", "spikes": ["BOL001","SPR001"]},
    "2025-01-02": {"event": "New Year’s Day (Day 2)", "spikes": ["BOL001"]},
    "2025-01-07": {"event": "Orthodox Christmas", "spikes": ["SPR001"]},
    "2025-02-15": {"event": "Statehood Day", "spikes": ["ASM001"]},
    "2025-02-16": {"event": "Statehood Day (Day 2)", "spikes": ["ASM001"]},
    "2025-02-17": {"event": "Statehood Day (Day 3)", "spikes": ["ASM001"]},
    "2025-04-18": {"event": "Good Friday", "spikes": ["PLT001"]},
    "2025-04-19": {"event": "Holy Saturday", "spikes": ["PLT001"]},
    "2025-04-20": {"event": "Easter Sunday", "spikes": ["ASM002"]},
    "2025-04-21": {"event": "Easter Monday", "spikes": ["ASM002"]},
    "2025-05-01": {"event": "Labor Day", "spikes": ["BOL001","NUT001"]},
    "2025-05-02": {"event": "Labor Day (Day 2)", "spikes": ["BOL001","NUT001"]},
    "2025-11-11": {"event": "Armistice Day", "spikes": ["SPR001"]},
    "2025-12-25": {"event": "Western Christmas", "spikes": ["BOL001","NUT001"]},
    "2026-01-01": {"event": "New Year’s Day", "spikes": ["BOL001","SPR001"]},
    "2026-01-02": {"event": "New Year’s Day (Day 2)", "spikes": ["BOL001"]},
    "2026-01-07": {"event": "Orthodox Christmas", "spikes": ["SPR001"]},
    "2026-02-15": {"event": "Statehood Day", "spikes": ["ASM001"]},
    "2026-02-16": {"event": "Statehood Day (Day 2)", "spikes": ["ASM001"]},
    "2026-02-17": {"event": "Statehood Day (Day 3)", "spikes": ["ASM001"]},
    "2026-04-10": {"event": "Good Friday", "spikes": ["PLT001"]},
    "2026-04-11": {"event": "Holy Saturday", "spikes": ["PLT001"]},
    "2026-04-12": {"event": "Easter Sunday", "spikes": ["ASM002"]},
    "2026-04-13": {"event": "Easter Monday", "spikes": ["ASM002"]},
    "2026-05-01": {"event": "Labor Day", "spikes": ["BOL001","NUT001"]},
    "2026-05-02": {"event": "Labor Day (Day 2)", "spikes": ["BOL001","NUT001"]},
    "2026-11-11": {"event": "Armistice Day", "spikes": ["SPR001"]},
    "2026-12-25": {"event": "Western Christmas", "spikes": ["BOL001","NUT001"]},
}

start_date = datetime(2025, 6, 1)
end_date = datetime(2026, 6, 30)
workdays_per_week = [0,1,2,3,4]  # Mon–Fri

# --- Export folder ---
export_dir = "Export"
os.makedirs(export_dir, exist_ok=True)

# --- Generate daily data ---
rows = []
stock_levels = {code: 1000 for code in items}  # initial stock
current_date = start_date
while current_date <= end_date:
    if current_date.weekday() in workdays_per_week:
        for code, meta in items.items():
            consumed = int(np.random.normal(meta["avg_daily"], meta["avg_daily"]*0.15))
            consumed = max(consumed, 0)
            wasted = int(consumed * meta["waste_pct"] + np.random.randint(0,3))

            order_placed = "Yes" if current_date.weekday() == 0 else "No"
            order_qty = None
            arrivals = 0

            # Order logic with reorder point + max stock
            if order_placed == "Yes":
                if stock_levels[code] <= meta["reorder_point"]:
                    need = meta["avg_daily"] * meta["lead_time"]
                    order_qty = int(np.ceil(need * (1 + meta["waste_pct"]) / meta["moq"]) * meta["moq"])
                    order_qty = min(order_qty, meta["max_stock"] - stock_levels[code])
                    arrivals = order_qty
                    stock_levels[code] += arrivals

            # Update stock with shortage handling
            start_stock = stock_levels[code]
            total_demand = consumed + wasted

            if total_demand > start_stock:
                # Not enough stock → cap consumption and record shortage
                shortage = total_demand - start_stock
                consumed = max(start_stock - wasted, 0)  # adjust consumed so wasted still counts
                wasted = max(start_stock - consumed, 0) # adjust wasted if needed
                stock_levels[code] = 0
            else:
                shortage = 0
                stock_levels[code] -= total_demand

            end_stock = stock_levels[code]

            # Contextual factors
            season = ("Winter" if current_date.month in [12,1,2] else
                      "Spring" if current_date.month in [3,4,5] else
                      "Summer" if current_date.month in [6,7,8] else
                      "Autumn")
            holiday_event = None
            demand_spike = "No"
            notes = None
            if current_date.strftime("%Y-%m-%d") in holidays:
                holiday_event = holidays[current_date.strftime("%Y-%m-%d")]["event"]
                if code in holidays[current_date.strftime("%Y-%m-%d")]["spikes"]:
                    demand_spike = "Yes"
                    notes = "Pre-holiday stocking"
                else:
                    notes = "Normal demand"

            rows.append({
                "Date": current_date,
                "Week": current_date.isocalendar()[1],
                "Month": current_date.month,
                "Year": current_date.year,
                "Item Code": code,
                "Item Name": meta["name"],
                "Consumed": consumed,
                "Wasted": wasted,
                "Shortage": shortage,   # NEW COLUMN
                "Order Placed": order_placed,
                "Order Qty": order_qty,
                "Start Stock": start_stock,
                "Arrivals": arrivals,
                "End Stock": end_stock,
                "Max Stock": meta["max_stock"],
                "Reorder Point": meta["reorder_point"],
                "Season": season,
                "Holiday/Event": holiday_event,
                "Demand Spike": demand_spike,
                "Notes": notes
            })
    current_date += timedelta(days=1)


df = pd.DataFrame(rows)

# --- Export function with Month_Year sheet names ---
def export_monthly(df, filename, columns, merge_week=True):
    wb = Workbook()
    wb.remove(wb.active)
    for (year, month) in sorted(df[["Year","Month"]].drop_duplicates().values.tolist()):
        ws = wb.create_sheet(title=f"{calendar.month_name[month]}_{year}")
        month_df = df[(df["Year"]==year) & (df["Month"]==month)][columns]
        for r in dataframe_to_rows(month_df, index=False, header=True):
            ws.append(r)
        if merge_week:
            week_col = 1
            current_week = None
            start_row = None
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                week_val = row[week_col-1].value
                if week_val != current_week:
                    if current_week is not None:
                        ws.merge_cells(start_row=start_row, start_column=week_col,
                                       end_row=i-1, end_column=week_col)
                        ws.cell(start_row, week_col).alignment = Alignment(vertical="center")
                    current_week = week_val
                    start_row = i
            if current_week is not None:
                ws.merge_cells(start_row=start_row, start_column=week_col,
                               end_row=ws.max_row, end_column=week_col)
                ws.cell(start_row, week_col).alignment = Alignment(vertical="center")
    wb.save(filename)

# --- Export files ---
export_monthly(df,
               os.path.join(export_dir, "Orders.xlsx"),
               ["Week","Date","Item Code","Order Qty","Order Placed"])

export_monthly(df,
               os.path.join(export_dir, "Storage.xlsx"),
               ["Week","Date","Item Code","Start Stock","Consumed","Wasted","Shortage","Arrivals","End Stock","Max Stock","Reorder Point"])

export_monthly(df,
               os.path.join(export_dir, "Context.xlsx"),
               ["Week","Date","Item Code","Season","Holiday/Event","Demand Spike","Notes"])

# --- Export Master Data ---
master_data = master_df.copy()
master_data.to_excel(os.path.join(export_dir, "MasterData.xlsx"), index=False)

